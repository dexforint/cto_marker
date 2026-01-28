"""
Провайдер для обработки электронных таблиц (XLSX, CSV).

Модуль предоставляет SpreadSheetProvider, который конвертирует электронные
таблицы в формат PDF для последующей обработки единообразным способом.
Процесс включает чтение XLSX файла, преобразование в HTML таблицу
с сохранением структуры и объединенных ячеек, и конвертацию в PDF.

Автор: Marker Team
"""

import os
import tempfile

from marker.providers.pdf import PdfProvider

# CSS стили для отображения таблиц в альбомной ориентации
css = '''
@page {
    size: A4 landscape;  # Альбомная ориентация страницы A4
    margin: 1.5cm;  # Отступы по 1.5 см со всех сторон
}

table {
    width: 100%;  # Ширина таблицы 100%
    border-collapse: collapse;  # Слияние границ ячеек
    break-inside: auto;  # Автоматический перенос таблиц
    font-size: 10pt;  # Размер шрифта 10 пунктов
}

tr {
    break-inside: avoid;  # Избегать переноса строк таблицы
    page-break-inside: avoid;  # Избегать разрыва страницы внутри строки
}

td {
    border: 0.75pt solid #000;  # Границы ячеек: 0.75pt, сплошные, черные
    padding: 6pt;  # Внутренние отступы ячеек
}
'''


class SpreadSheetProvider(PdfProvider):
    """
    Провайдер для обработки электронных таблиц (XLSX, CSV).
    
    Наследуется от PdfProvider и обеспечивает конвертацию электронных
    таблиц в промежуточный формат PDF для последующей обработки.
    
    Процесс обработки:
    1. Создание временного PDF файла
    2. Чтение XLSX файла с помощью openpyxl
    3. Преобразование каждого листа в HTML таблицу
    4. Сохранение структуры и объединенных ячеек
    5. Конвертация HTML в PDF
    6. Инициализация родительского PdfProvider с временным PDF
    """
    def __init__(self, filepath: str, config=None):
        """
        Инициализация провайдера электронных таблиц.
        
        Args:
            filepath (str): Путь к XLSX/CSV файлу
            config: Конфигурация провайдера (опционально)
        """
        # Создаем временный PDF файл для промежуточного хранения
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=f".pdf")
        self.temp_pdf_path = temp_pdf.name
        temp_pdf.close()

        # Конвертируем XLSX в PDF
        try:
            self.convert_xlsx_to_pdf(filepath)
        except Exception as e:
            # В случае ошибки конвертации удаляем временный файл и выбрасываем исключение
            raise RuntimeError(f"Failed to convert {filepath} to PDF: {e}")

        # Инициализируем родительский PdfProvider с временным PDF файлом
        super().__init__(self.temp_pdf_path, config)

    def __del__(self):
        """
        Деструктор для очистки временных файлов.
        
        Удаляет временный PDF файл при уничтожении объекта.
        """
        if os.path.exists(self.temp_pdf_path):
            os.remove(self.temp_pdf_path)

    def convert_xlsx_to_pdf(self, filepath: str):
        """
        Конвертирует XLSX файл в PDF формат.
        
        Процесс включает:
        1. Чтение XLSX файла с помощью openpyxl
        2. Обработку каждого листа как отдельной HTML секции
        3. Преобразование таблиц в HTML с сохранением структуры
        4. Конвертацию HTML в PDF с альбомной ориентацией
        
        Args:
            filepath (str): Путь к исходному XLSX файлу
        """
        from weasyprint import CSS, HTML
        from openpyxl import load_workbook

        html = ""
        # Загружаем рабочую книгу Excel
        workbook = load_workbook(filepath)
        if workbook is not None:
            # Обрабатываем каждый лист
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Создаем HTML секцию для каждого листа с заголовком и таблицей
                html += f'<div><h1>{sheet_name}</h1>' + self._excel_to_html_table(sheet) + '</div>'
        else:
            # Если не удалось загрузить файл, выбрасываем исключение
            raise ValueError("Invalid XLSX file")

        # Конвертируем HTML в PDF с применением стилей
        HTML(string=html).write_pdf(
            self.temp_pdf_path,
            stylesheets=[CSS(string=css), self.get_font_css()]
        )

    @staticmethod
    def _get_merged_cell_ranges(sheet):
        """
        Извлекает информацию об объединенных ячейках на листе.
        
        Создает словарь с информацией о позиции, размерах и границах
        объединенных ячеек для последующего корректного отображения в HTML.
        
        Args:
            sheet: Объект листа из openpyxl
            
        Returns:
            dict: Словарь с информацией об объединенных ячейках
        """
        merged_info = {}
        # Проходим по всем диапазонам объединенных ячеек
        for merged_range in sheet.merged_cells.ranges:
            # Получаем границы диапазона
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_info[(min_row, min_col)] = {
                'rowspan': max_row - min_row + 1,  # Количество строк в объединении
                'colspan': max_col - min_col + 1,  # Количество столбцов в объединении
                'range': merged_range  # Оригинальный диапазон
            }
        return merged_info

    def _excel_to_html_table(self, sheet):
        """
        Преобразует лист Excel в HTML таблицу.
        
        Обрабатывает каждую ячейку листа, создавая HTML таблицу с правильными
        rowspan и colspan атрибутами для объединенных ячеек.
        
        Args:
            sheet: Объект листа из openpyxl
            
        Returns:
            str: HTML строка с таблицей
        """
        # Получаем информацию об объединенных ячейках
        merged_cells = self._get_merged_cell_ranges(sheet)

        # Начинаем создавать HTML таблицу
        html = f'<table>'

        # Множество для отслеживания ячеек, которые нужно пропустить
        skip_cells = set()

        # Обрабатываем каждую строку
        for row_idx, row in enumerate(sheet.rows, 1):
            html += '<tr>'
            # Обрабатываем каждую ячейку в строке
            for col_idx, cell in enumerate(row, 1):
                # Пропускаем ячейки, которые являются частью объединенного диапазона
                if (row_idx, col_idx) in skip_cells:
                    continue

                # Проверяем, является ли эта ячейка началом объединенного диапазона
                merge_info = merged_cells.get((row_idx, col_idx))
                if merge_info:
                    # Добавляем ячейки в список для пропуска
                    for r in range(row_idx, row_idx + merge_info['rowspan']):
                        for c in range(col_idx, col_idx + merge_info['colspan']):
                            if (r, c) != (row_idx, col_idx):
                                skip_cells.add((r, c))

                    # Добавляем объединенную ячейку с rowspan и colspan
                    value = cell.value if cell.value is not None else ''
                    html += f'<td rowspan="{merge_info["rowspan"]}" colspan="{merge_info["colspan"]}">{value}'
                else:
                    # Обычная ячейка
                    value = cell.value if cell.value is not None else ''
                    html += f'<td>{value}'

                html += '</td>'
            html += '</tr>'
        html += '</table>'
        return html
